import requests
from bs4 import BeautifulSoup
import openpyxl

# URL do site
login_url = 'https://ava3.cs.edu.br/login/index.php'

# Dados de autenticação
usuario = 'Teste'
senha = 'Teste'

# Texto a ser verificado no HTML
texto_especifico = "Seção 1"

# Criar uma sessão
session = requests.Session()

# Realizar a solicitação GET para obter o token de autenticação
response = session.get(login_url)
soup = BeautifulSoup(response.content, 'html.parser')

# Encontrar o token de autenticação (se existir)
token = soup.find('input', {'name': 'logintoken'})['value']

# Dados para a solicitação POST de autenticação
login_data = {
    'username': usuario,
    'password': senha,
    'logintoken': token
}

# Realizar a solicitação POST para fazer o login
login_response = session.post(login_url, data=login_data)

# Criar uma nova planilha para os resultados
resultado_workbook = openpyxl.Workbook()
resultado_sheet = resultado_workbook.active
resultado_sheet.append(['Link', 'Contém o Texto'])

# Verificar se o login foi bem-sucedido
if 'Invalid login' not in login_response.text:
    print("Login bem-sucedido!")

    # Ler os URLs da planilha "links.xlsx"
    workbook = openpyxl.load_workbook('link.xlsx')
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        link = row[0]  # Assumindo que o link está na primeira coluna
        # Acessar cada link após o login bem-sucedido
        link_response = session.get(link)

        # Verificar se o texto específico está presente no HTML
        if texto_especifico in link_response.text:
            contem_texto = "Sim"
        else:
            contem_texto = "Não"

        # Adicionar os resultados à nova planilha
        resultado_sheet.append([link, contem_texto])

# Salvar a nova planilha com os resultados
resultado_workbook.save('resultados.xlsx')

# Certifique-se de encerrar a sessão quando terminar
session.close()
